import django_filters
from rest_framework.views import APIView
from rest_framework import viewsets, status
from .models import Venue, Event, EventImage
from .serializers import VenueSerializer, EventSerializer
from .filters import EventFilter
from rest_framework.filters import SearchFilter, OrderingFilter
from .permissions import IsAdminReadOnly
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from .pagination import TestPagination
from openpyxl import load_workbook, Workbook
from datetime import datetime
from django.http import HttpResponse
from rest_framework.response import Response
from .filters import EventFilter

class VenueViewSet(viewsets.ModelViewSet):
    queryset = Venue.objects.all()
    serializer_class = VenueSerializer

class EventViewSet(viewsets.ModelViewSet):
    serializer_class = EventSerializer
    permission_classes = [IsAdminReadOnly]
    pagination_class = TestPagination

    filter_backends = (
        django_filters.rest_framework.DjangoFilterBackend,
        SearchFilter,
        OrderingFilter,
    )

    filterset_class = EventFilter
    search_fields = ['title', 'venue__name']
    ordering_fields = ['start_datetime', 'end_datetime', 'title']
    ordering = ['title']

    def get_queryset(self):
        queryset = Event.objects.all()

        if not self.request.user.is_superuser:
            queryset = queryset.filter(status='PUBLISHED')

        return queryset

    def perform_create(self, serializer):
        serializer.save(author=self.request.user)

class UploadEventImagesAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        event_id = request.data.get('event_id')
        event = Event.objects.filter(id=event_id).first()

        if not event:
            return Response({"error": "Событие не найдено"}, status=status.HTTP_404_NOT_FOUND)
        
        images = request.FILES.get('images')

        if not images:
            return Response({"error": "Изображения не переданы"}, status=status.HTTP_400_BAD_REQUEST)
        
        event_images = []
        for image in images:
            event_image = EventImage.objects.create(event=event, image=image)
            event_images.append(event_image)

        if event_images:
            first_image = event_image[0]

            if not first_image.preview:
                first_image.save()
        return Response({
            'message':'Изображения успешно загружены',
            'images': [image.id for image in event_images]
        }, status=status.HTTP_201_CREATED)
        

class ImportEventsAPIView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')

        if not file:
            return Response({"error":"Файл не передан"}, status=400)
        
        if not file.name.endswith('.xlsx'):
            return Response({'detail':'ожидается файл типа xlsx'}, status=400)
        
        try:
            workbook = load_workbook(file)
            sheet = workbook.active
        except Exception as e:
            return Response({'detail':f'ошибка чтения файла: {str(e)}'})
        
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            return Response({'detail':'файл пуст'}, status=400)
        
        data_rows = rows[1:]

        created_count = 0
        errors = []

        for row_index, row in enumerate(data_rows, start=2):
            try:
                if len(row) < 9:
                    errors.append({
                        'row': row_index,
                        'error': f'Недостаточно колонок в строке (ожидается 9, получено {len(row)})'
                    })
                    continue
                title = row[0]
                desk = row[1]
                publish_datetime = row[2]
                start_datetime = row[3]
                end_datetime = row[4]
                venue_name = row[5]
                latitude, longitude = map(float, row[6].split(','))
                rating = int(row[7] or 0)
                weather = row[8]

                if not title or not venue_name:
                    continue

                try:
                    if rating is None or rating == '':
                        rating = 0
                    else:
                        rating = int(rating)
                except ValueError:
                    rating = 0

                if rating < 0: 
                    rating = 0
                if rating > 25: 
                    rating = 25



                try:
                    if not isinstance(publish_datetime, datetime):
                        publish_datetime = datetime.strptime(str(publish_datetime), '%Y-%m-%d %H:%M')
                    if not isinstance(start_datetime, datetime):
                        start_datetime = datetime.strptime(str(start_datetime), '%Y-%m-%d %H:%M')
                    if not isinstance(end_datetime, datetime):
                        end_datetime = datetime.strptime(str(end_datetime), '%Y-%m-%d %H:%M')
                except ValueError as e:
                    errors.append({
                        'row': row_index,
                        'error': f'Неверный формат даты: {str(e)}'
                    })
                    continue


                venue, _ = Venue.objects.get_or_create(
                    name=venue_name,
                    defaults={
                        'latitude': latitude,
                        'longitude': longitude,
                        'weather':weather,
                    },
                )

                if Event.objects.filter(title=title, start_datetime=start_datetime, venue=venue).exists():
                    continue


                Event.objects.create(
                    title=title,
                    desk=desk,
                    publish_datetime=publish_datetime,
                    start_datetime=start_datetime,
                    end_datetime=end_datetime,
                    author=request.user,
                    venue=venue,
                    rating=rating,
                    status="DRAFT",
                )

                created_count += 1

            except Exception as e:
                errors.append({
                    'row': row_index,
                    'error': str(e)
                })
        return Response({
            'created': created_count,
            'errors':errors
        },
        status=201 if not errors else 207
    )
        
class ExportEventsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        queryset = Event.objects.all()
        filtered_queryset = EventFilter(request.GET, queryset=queryset).qs

        workbook = Workbook()
        ws = workbook.active
        ws.title = "Events"

        headers = [
            "title",
            "desk",
            "publish_datetime",
            "start_datetime",
            "end_datetime",
            "venue_name",
            "latitude",
            "longitude",
            'weather',
            "rating",
            "status"
        ]
        ws.append(headers)

        for event in filtered_queryset:
            ws.append([
                event.title,
                event.desk,
                event.publish_datetime.strftime('%Y-%m-%d %H:%M'),
                event.start_datetime.strftime('%Y-%m-%d %H:%M'),
                event.end_datetime.strftime('%Y-%m-%d %H:%M'),
                event.venue.name if event.venue else "",
                event.venue.latitude if event.venue else "",
                event.venue.longitude if event.venue else "",
                event.rating,
                event.status,
                event.venue.weather if event.venue else ""
            ])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)

            response['Content-Disposition'] = 'attachment; filename=events_export.xlsx'
            workbook.save(response)
        return response
        
# /api/events/export-xlsx/?start_datetime_after=2026-01-01&start_datetime_before=2026-02-01&venue__id=1&rating_min=10&rating_max=25

#где:
# start_datetime_after / start_datetime_before — диапазон начала события

# end_datetime_after / end_datetime_before — диапазон окончания

# publish_datetime_after / publish_datetime_before — диапазон публикации

# venue__id — фильтр по месту проведения

# rating_min / rating_max — диапазон рейтинга
        

        
